import csv
import json
import shutil
import uuid
from datetime import date
from pathlib import Path
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, UploadFile, status
from sqlalchemy.orm import Session

from app.api.deps import get_business_profile_id
from app.audit import record_audit
from app.config import get_settings
from app.database import get_db
from app.idempotency import begin_idempotent_request, complete_idempotent_request
from app.models import Category, Product, ProductPrice, Supplier, UploadedFile
from app.product_identifiers import (
    ProductIdentifierError,
    assert_barcode_available,
    generate_product_identifiers,
    normalize_manual_barcode,
)
from app.product_identity import clean_product_name, find_active_product_by_name
from app.schemas import FileProductImportResult, FileProductImportSubmit, ProductOut, UploadedFileOut
from app.services import product_metrics, record_product_quantity, retry_on_deadlock

router = APIRouter(prefix="/files", tags=["Files"])
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

ALLOWED_EXTENSIONS = {".csv", ".xls", ".xlsx"}
PREVIEW_LIMIT = 100

PRODUCT_NAME_KEYS = {"product", "productname", "item", "itemname", "name", "description", "descriptionofgoods", "material"}
SKU_KEYS = {"sku", "code", "productcode", "itemcode"}
BARCODE_KEYS = {"barcode", "barcodeno", "barcodenumber", "ean", "upc"}
CATEGORY_KEYS = {"category", "type", "group"}
SUPPLIER_KEYS = {"supplier", "suppliername", "vendor", "vendorname"}
QUANTITY_KEYS = {"quantity", "qty", "stock", "newstock", "newqty", "qtybought", "boughtqty", "boughtquantity", "availableqty"}
MRP_KEYS = {"mrp", "maximumretailprice"}
BUY_PRICE_KEYS = {"buyprice", "purchaseprice", "costprice", "cost", "boughtprice", "landingprice"}
SELL_PRICE_KEYS = {"sellprice", "sellingprice", "saleprice", "price", "rate"}
GST_KEYS = {"gst", "gstrate", "tax", "taxrate"}
UNIT_KEYS = {"unit", "unitlabel", "uom", "format"}


def _upload_root() -> Path:
    settings = get_settings()
    root = Path(settings.upload_dir)
    if not root.is_absolute():
        root = Path(__file__).resolve().parents[2] / root
    return root / "files"


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_to_preview(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, str]], int]:
    if not rows:
        return [], [], 0

    headers = []
    seen_headers: dict[str, int] = {}
    for index, value in enumerate(rows[0]):
        base_header = _clean_cell(value) or f"Column {index + 1}"
        seen_headers[base_header] = seen_headers.get(base_header, 0) + 1
        headers.append(base_header if seen_headers[base_header] == 1 else f"{base_header} {seen_headers[base_header]}")
    preview_rows = []
    for row in rows[1 : PREVIEW_LIMIT + 1]:
        preview_rows.append(
            {
                headers[index] if index < len(headers) else f"Column {index + 1}": _clean_cell(value)
                for index, value in enumerate(row)
            }
        )
    return headers, preview_rows, max(0, len(rows) - 1)


def _rows_to_dicts(rows: list[list[Any]]) -> list[dict[str, str]]:
    if not rows:
        return []
    headers = []
    seen_headers: dict[str, int] = {}
    for index, value in enumerate(rows[0]):
        base_header = _clean_cell(value) or f"Column {index + 1}"
        seen_headers[base_header] = seen_headers.get(base_header, 0) + 1
        headers.append(base_header if seen_headers[base_header] == 1 else f"{base_header} {seen_headers[base_header]}")
    return [
        {
            headers[index] if index < len(headers) else f"Column {index + 1}": _clean_cell(value)
            for index, value in enumerate(row)
        }
        for row in rows[1:]
    ]


def _extract_csv(path: Path) -> tuple[list[str], list[dict[str, str]], int]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        return _rows_to_preview(list(reader))


def _extract_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]], int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required to read XLSX files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return _rows_to_preview(rows)


def _extract_xls(path: Path) -> tuple[list[str], list[dict[str, str]], int]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="xlrd is required to read XLS files") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
    return _rows_to_preview(rows)


def _extract_file(path: Path, extension: str) -> tuple[list[str], list[dict[str, str]], int]:
    if extension == ".csv":
        return _extract_csv(path)
    if extension == ".xlsx":
        return _extract_xlsx(path)
    return _extract_xls(path)


def _all_rows_from_path(path: Path, extension: str) -> list[dict[str, str]]:
    if extension == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return [{key: _clean_cell(value) for key, value in row.items()} for row in reader]
    if extension == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return _rows_to_dicts(raw_rows)
    import xlrd

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    raw_rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
    return _rows_to_dicts(raw_rows)


def _normalize_key(value: str) -> str:
    return "".join(character for character in str(value or "").lower() if character.isalnum())


def _get_value(row: dict[str, Any], keys: set[str]) -> str:
    normalized = {_normalize_key(key): value for key, value in row.items()}
    for key in keys:
        if key in normalized and str(normalized[key]).strip():
            return str(normalized[key]).strip()
    return ""


def _decimal_value(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    clean_value = str(value).replace(",", "").replace("₹", "").strip()
    if not clean_value:
        return default
    try:
        return Decimal(clean_value)
    except (InvalidOperation, ValueError):
        return default


def _ensure_category(db: Session, name: str | None, product_name: str = "") -> Category:
    category_name = (name or "").strip()
    if not category_name:
        existing_categories = db.query(Category).filter(Category.is_active.is_(True)).order_by(Category.name.asc()).all()
        product_blob = _normalize_key(product_name)
        for category in existing_categories:
            normalized_category = _normalize_key(category.name)
            if normalized_category and normalized_category in product_blob:
                return category
        if existing_categories:
            return existing_categories[0]
        category_name = "General"
    category = db.query(Category).filter(Category.name.ilike(category_name)).first()
    if category:
        return category
    category = Category(name=category_name, is_active=True)
    db.add(category)
    db.flush()
    return category


def _ensure_supplier(db: Session, name: str | None, business_profile_id: int | None) -> Supplier:
    supplier_name = (name or "Unknown").strip() or "Unknown"
    query = db.query(Supplier).filter(Supplier.name.ilike(supplier_name))
    if business_profile_id is not None:
        query = query.filter(Supplier.business_profile_id == business_profile_id)
    supplier = query.first()
    if supplier:
        return supplier
    supplier = Supplier(name=supplier_name, business_profile_id=business_profile_id, is_active=True)
    db.add(supplier)
    db.flush()
    return supplier


def _product_out(product: Product) -> ProductOut:
    return ProductOut.model_validate(
        {
            **product.__dict__,
            **product_metrics(product),
            "quantity_history": getattr(product, "quantities", []),
            "qualities": getattr(product, "qualities", []),
            "price_history": getattr(product, "price_history", []),
            "discounts": getattr(product, "discounts", []),
        }
    )


def _all_rows_from_file(file_record: UploadedFile) -> list[dict[str, str]]:
    if file_record.rows_json:
        try:
            rows = json.loads(file_record.rows_json)
            if isinstance(rows, list):
                return [{str(key): _clean_cell(value) for key, value in row.items()} for row in rows if isinstance(row, dict)]
        except json.JSONDecodeError:
            pass

    path = Path(file_record.file_path)
    extension = f".{file_record.file_type}"
    if not path.exists():
        try:
            preview_rows = json.loads(file_record.preview_json or "[]")
            if isinstance(preview_rows, list) and preview_rows:
                return [
                    {str(key): _clean_cell(value) for key, value in row.items()}
                    for row in preview_rows
                    if isinstance(row, dict)
                ]
        except json.JSONDecodeError:
            pass
        raise HTTPException(
            status_code=400,
            detail="Uploaded source file is missing on the server and no saved preview rows are available. Please upload the file again and submit.",
        )
    return _all_rows_from_path(path, extension)


def _file_to_schema(file_record: UploadedFile) -> UploadedFileOut:
    columns = json.loads(file_record.columns_json or "[]")
    preview_rows = json.loads(file_record.preview_json or "[]")
    return UploadedFileOut(
        id=file_record.id,
        business_profile_id=file_record.business_profile_id,
        original_name=file_record.original_name,
        stored_name=file_record.stored_name,
        file_url=file_record.file_url,
        file_type=file_record.file_type,
        row_count=file_record.row_count,
        columns=columns,
        preview_rows=preview_rows,
        created_at=file_record.created_at,
        updated_at=file_record.updated_at,
    )


@router.get("", response_model=list[UploadedFileOut])
def list_uploaded_files(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=100),
    business_profile_id: int | None = Depends(get_business_profile_id),
    db: Session = Depends(get_db),
) -> list[UploadedFileOut]:
    query = db.query(UploadedFile).filter(UploadedFile.is_active.is_(True)).order_by(UploadedFile.created_at.desc(), UploadedFile.id.desc())
    if business_profile_id is not None:
        query = query.filter(UploadedFile.business_profile_id == business_profile_id)
    return [_file_to_schema(file_record) for file_record in query.offset(skip).limit(limit).all()]


@router.post("", response_model=UploadedFileOut, status_code=status.HTTP_201_CREATED)
def upload_file(
    upload: UploadFile = File(...),
    business_profile_id: int | None = Depends(get_business_profile_id),
    db: Session = Depends(get_db),
) -> UploadedFileOut:
    if upload.size is not None and upload.size > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File size must not exceed 10 MB")
    extension = Path(upload.filename or "").suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only CSV, XLS, and XLSX files are allowed")

    upload_dir = _upload_root()
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(upload.filename or f"erp-file{extension}").name
    stored_name = f"{uuid.uuid4().hex[:12]}-{safe_name}"
    destination = upload_dir / stored_name
    with destination.open("wb") as output:
        shutil.copyfileobj(upload.file, output)

    columns, preview_rows, row_count = _extract_file(destination, extension)
    all_rows = _all_rows_from_path(destination, extension)
    file_record = UploadedFile(
        business_profile_id=business_profile_id,
        original_name=safe_name,
        stored_name=stored_name,
        file_url=f"/uploads/files/{stored_name}",
        file_path=str(destination),
        file_type=extension.lstrip("."),
        row_count=row_count,
        columns_json=json.dumps(columns),
        preview_json=json.dumps(preview_rows),
        rows_json=json.dumps(all_rows),
    )
    db.add(file_record)
    db.flush()
    record_audit(
        db,
        action="upload_file",
        entity_type="uploaded_file",
        entity_id=file_record.id,
        details={"fileName": safe_name, "rowCount": row_count},
    )
    db.commit()
    db.refresh(file_record)
    return _file_to_schema(file_record)


@router.post("/{file_id}/submit-products", response_model=FileProductImportResult)
@retry_on_deadlock()
def submit_file_products(
    file_id: int,
    payload: FileProductImportSubmit,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    business_profile_id: int | None = Depends(get_business_profile_id),
    db: Session = Depends(get_db),
) -> FileProductImportResult:
    idem = begin_idempotent_request(
        db,
        idempotency_key,
        f"ERP:POST:/files/{file_id}/submit-products",
        {"file_id": file_id, **payload.model_dump()},
    )
    if idem.replay_body is not None:
        return idem.replay_body
    file_record = db.get(UploadedFile, file_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="Uploaded file not found")
    if business_profile_id is not None and file_record.business_profile_id not in {None, business_profile_id}:
        raise HTTPException(status_code=404, detail="Uploaded file not found")

    # `rows` is retained for older clients. The web preview only contains the
    # first 100 rows, so edits now arrive as overrides and the complete saved
    # file remains the import source.
    rows = payload.rows if payload.rows is not None else _all_rows_from_file(file_record)
    if payload.row_overrides is not None:
        rows = _all_rows_from_file(file_record)
        for row_index, override in enumerate(payload.row_overrides):
            if row_index < len(rows) and isinstance(override, dict):
                rows[row_index] = {str(key): _clean_cell(value) for key, value in override.items()}
    result = FileProductImportResult()
    changed_products: list[Product] = []

    indexed_rows = sorted(
        enumerate(rows, start=1),
        key=lambda item: (
            (_get_value(item[1], SKU_KEYS) or "").strip().lower(),
            (_get_value(item[1], PRODUCT_NAME_KEYS) or "").strip().lower(),
            item[0],
        ),
    )

    for row_index, row in indexed_rows:
        product_name = clean_product_name(_get_value(row, PRODUCT_NAME_KEYS))
        sku = _get_value(row, SKU_KEYS)
        barcode = _get_value(row, BARCODE_KEYS)
        if not product_name and not sku:
            result.skipped += 1
            result.messages.append(f"Row {row_index}: skipped because product name/code is empty")
            continue

        quantity = _decimal_value(_get_value(row, QUANTITY_KEYS), Decimal("0"))
        mrp = _decimal_value(_get_value(row, MRP_KEYS), Decimal("0"))
        buy_price = _decimal_value(_get_value(row, BUY_PRICE_KEYS), Decimal("0"))
        sell_price = _decimal_value(_get_value(row, SELL_PRICE_KEYS), Decimal("0"))
        if sell_price <= 0 and mrp > 0:
            sell_price = mrp
        if mrp <= 0 and sell_price > 0:
            mrp = sell_price
        if buy_price <= 0 and sell_price > 0:
            buy_price = sell_price
        if mrp <= 0:
            mrp = Decimal("0.01")
        if buy_price < 0:
            buy_price = Decimal("0")
        if sell_price < 0:
            sell_price = Decimal("0")
        gst_rate = _decimal_value(_get_value(row, GST_KEYS), Decimal("18"))
        unit_label = _get_value(row, UNIT_KEYS) or "Pieces"
        category = _ensure_category(db, _get_value(row, CATEGORY_KEYS), product_name)
        supplier = _ensure_supplier(db, _get_value(row, SUPPLIER_KEYS), business_profile_id)

        product_query = db.query(Product)
        if business_profile_id is not None:
            product_query = product_query.filter(Product.business_profile_id == business_profile_id)
        product = None
        if sku:
            product = product_query.filter(Product.sku.ilike(sku)).with_for_update().first()
        if product is None and product_name:
            product = find_active_product_by_name(
                db,
                name=product_name,
                business_profile_id=business_profile_id,
                lock=True,
            )

        if product:
            try:
                manual_barcode = normalize_manual_barcode(barcode)
                if manual_barcode:
                    assert_barcode_available(
                        db,
                        barcode=manual_barcode,
                        business_profile_id=business_profile_id,
                        exclude_product_id=product.id,
                    )
                    product.barcode = manual_barcode
            except ProductIdentifierError as exc:
                result.skipped += 1
                result.messages.append(f"Row {row_index}: {exc.message}")
                continue
            old_stock = Decimal(product.qty_bought)
            old_sold = Decimal(product.qty_sold)
            old_mrp = Decimal(product.mrp)
            old_buy_price = Decimal(product.buy_price)
            old_sell_price = Decimal(product.sell_price)
            product.qty_bought += quantity
            product.category_id = category.id
            product.category = category.name
            product.supplier_id = supplier.id
            product.supplier = supplier.name
            product.unit_label = unit_label
            product.unit_type = unit_label.lower().replace(" ", "_") or "pieces"
            if mrp > Decimal("0.01"):
                product.mrp = mrp
            if buy_price > 0:
                product.buy_price = buy_price
            if sell_price > 0:
                product.sell_price = sell_price
            product.gst_rate = gst_rate
            product.reorder_level = max(Decimal("0"), Decimal(product.qty_sold))
            record_product_quantity(
                db,
                product,
                transaction_type="file_import_stock",
                quantity_change=quantity,
                old_stock=old_stock,
                new_stock=Decimal(product.qty_bought),
                sold_stock=Decimal(product.qty_sold),
                note=f"Imported from {file_record.original_name}",
            )
            if (
                Decimal(product.mrp) != old_mrp
                or Decimal(product.buy_price) != old_buy_price
                or Decimal(product.sell_price) != old_sell_price
            ):
                db.add(
                    ProductPrice(
                        product_id=product.id,
                        business_profile_id=business_profile_id,
                        effective_date=date.today(),
                        mrp=product.mrp,
                        buy_price=product.buy_price,
                        sell_price=product.sell_price,
                        source="file_import",
                        note=f"Imported from {file_record.original_name}",
                    )
                )
            result.updated += 1
            changed_products.append(product)
        else:
            try:
                manual_barcode = normalize_manual_barcode(barcode)
                identifiers = generate_product_identifiers(db, business_profile_id)
                if manual_barcode:
                    assert_barcode_available(
                        db,
                        barcode=manual_barcode,
                        business_profile_id=business_profile_id,
                    )
            except ProductIdentifierError as exc:
                result.skipped += 1
                result.messages.append(f"Row {row_index}: {exc.message}")
                continue
            product = Product(
                business_profile_id=business_profile_id,
                sku=identifiers.sku,
                barcode=manual_barcode or identifiers.barcode,
                name=product_name or sku,
                category_id=category.id,
                supplier_id=supplier.id,
                category=category.name,
                supplier=supplier.name,
                qty_bought=quantity,
                qty_sold=Decimal("0"),
                unit_type=unit_label.lower().replace(" ", "_") or "pieces",
                unit_label=unit_label,
                package_size=Decimal("1"),
                package_size_unit=unit_label,
                package_price=None,
                quantity_options=None,
                mrp=mrp,
                buy_price=buy_price,
                sell_price=sell_price,
                gst_rate=gst_rate,
                reorder_level=Decimal("0"),
                is_active=True,
            )
            db.add(product)
            db.flush()
            record_product_quantity(
                db,
                product,
                transaction_type="file_import_new_product",
                quantity_change=quantity,
                old_stock=Decimal("0"),
                new_stock=Decimal(product.qty_bought),
                sold_stock=Decimal("0"),
                note=f"Imported from {file_record.original_name}",
            )
            db.add(
                ProductPrice(
                    product_id=product.id,
                    business_profile_id=business_profile_id,
                    effective_date=date.today(),
                    mrp=product.mrp,
                    buy_price=product.buy_price,
                    sell_price=product.sell_price,
                    source="file_import",
                    note=f"Imported from {file_record.original_name}",
                )
            )
            result.created += 1
            changed_products.append(product)

    record_audit(
        db,
        action="submit_file_products",
        entity_type="uploaded_file",
        entity_id=file_record.id,
        details={"created": result.created, "updated": result.updated, "skipped": result.skipped},
    )
    db.flush()
    # Timestamp columns are database-generated. Refresh before serialising the
    # imported products so ProductOut always receives created_at/updated_at.
    for product in changed_products:
        db.refresh(product)
    result.products = [_product_out(product) for product in changed_products]
    if not result.messages:
        result.messages.append(f"Imported {result.created} new and updated {result.updated} products")
    complete_idempotent_request(idem, result, status.HTTP_200_OK)
    db.commit()
    return result


@router.delete("/{file_id}")
def delete_uploaded_file(
    file_id: int,
    business_profile_id: int = Depends(get_business_profile_id),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    file_record = (
        db.query(UploadedFile)
        .filter(
            UploadedFile.id == file_id,
            UploadedFile.business_profile_id == business_profile_id,
            UploadedFile.is_active.is_(True),
        )
        .first()
    )
    if not file_record:
        raise HTTPException(status_code=404, detail="Uploaded file not found")
    file_record.is_active = False
    record_audit(
        db,
        action="delete_file",
        entity_type="uploaded_file",
        entity_id=file_id,
        details={"fileName": file_record.original_name},
    )
    db.commit()
    return {"message": "File deleted"}
